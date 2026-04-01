#!/usr/bin/env python3
"""
Generate member-lookup.data.js from the Merged sheet in
Membership MiClub and GG.xlsx.

Outputs:
  tools/member-lookup/data.js
  data/exports/member-lookup.data.js
  site/assets/data/member-lookup.data.js
"""

import json, re, hashlib
from pathlib import Path
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed — run: pip3 install openpyxl")
    raise

BASE   = Path(__file__).parent
XLSX   = Path.home() / "Desktop" / "Golf Ops Data For Dashboard" / "Membership" / "Membership MiClub and GG.xlsx"
TODAY  = date.today()

OUTPUTS = [
    BASE / "tools"  / "member-lookup" / "data.js",
    BASE / "data"   / "exports"       / "member-lookup.data.js",
    BASE / "site"   / "assets" / "data" / "member-lookup.data.js",
]

# ---------------------------------------------------------------------------
# Column header → index mapping from Merged sheet
# ---------------------------------------------------------------------------
COLS = {
    "status": 0, "match_method": 1,
    "first": 2, "last": 3, "full": 4, "gender": 5,
    "member_code": 6, "member_id": 7, "membership_type": 8,
    "age": 9, "dob": 10, "join_date": 11, "handicap": 12,
    "mobile": 13, "email_mc": 14,
    "city": 15, "postcode": 16, "state": 17,
    "ggs_id": 18, "member_card_id": 19, "email_gg": 20,
    "birthday_gg": 21, "membership_date_gg": 22, "cell_phone_gg": 23,
    "driver": 24, "woods": 25, "irons": 26, "wedges": 27, "putter": 28,
    "ball_brand": 29, "glove_brand": 30, "glove_size": 31,
    "apparel_brand": 32, "apparel_size": 33,
    "shoe_brand": 34, "shoe_size": 35,
}

AGE_GROUPS  = [(0,17,"<18"),(18,29,"18–29"),(30,39,"30s"),(40,49,"40s"),
               (50,59,"50s"),(60,69,"60s"),(70,79,"70s"),(80,999,"80+")]

TENURE_BUCKETS = [(0,1,"New (<2yr)"),(2,4,"2–4yr"),(5,9,"5–9yr"),
                  (10,19,"10–19yr"),(20,49,"20–49yr"),(50,999,"50+yr")]


def clean(v):
    if v is None: return ""
    s = str(v).strip()
    return "" if s.lower() in ("none","n/a","#n/a","nan") else s


def age_group(age):
    if not age: return ""
    for lo, hi, label in AGE_GROUPS:
        if lo <= age <= hi: return label
    return "80+"


def tenure_bucket(yrs):
    if yrs is None: return ""
    for lo, hi, label in TENURE_BUCKETS:
        if lo <= yrs <= hi: return label
    return "50+yr"


def fmt_date(v):
    """Return DD/MM/YYYY string from various input types."""
    if not v: return ""
    if isinstance(v, (datetime, date)):
        try: return v.strftime("%d/%m/%Y")
        except: return ""
    s = clean(v)
    if not s: return ""
    # Already formatted
    if re.match(r"\d{2}/\d{2}/\d{4}", s): return s
    # YYYY-MM-DD
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m: return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    return s


def parse_year(v):
    s = fmt_date(v)
    if not s: return None
    try: return int(s.split("/")[2])
    except: return None


def make_id(first, last, ggs_id, member_code):
    if ggs_id: return str(ggs_id)
    key = f"{first}|{last}|{member_code}"
    return "mc_" + hashlib.md5(key.encode()).hexdigest()[:16]


def main():
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found")
        return

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    if "Merged" not in wb.sheetnames:
        print("ERROR: 'Merged' sheet not found — run the cross-reference script first")
        return

    ws = wb["Merged"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data    = rows[1:]
    print(f"Merged sheet: {len(data)} member rows")

    members = []

    for r in data:
        def g(col): return clean(r[COLS[col]])

        status = g("status")

        first  = g("first")
        last   = g("last")
        gender = g("gender")
        if gender in ("M","m"): gender = "Male"
        if gender in ("F","f"): gender = "Female"

        # Prefer MiClub email, fall back to GG
        email  = g("email_mc") or g("email_gg")
        phone  = g("mobile")   or g("cell_phone_gg")

        # Prefer MiClub dates; fall back to GG
        dob_raw  = r[COLS["dob"]]       or r[COLS["birthday_gg"]]
        join_raw = r[COLS["join_date"]] or r[COLS["membership_date_gg"]]

        birthday = fmt_date(dob_raw)
        join_date = fmt_date(join_raw)
        join_year = parse_year(join_raw)

        # Age from MiClub (already computed), or derive from birthday
        age = None
        mc_age = r[COLS["age"]]
        try:
            if mc_age and str(mc_age).replace('.','',1).isdigit():
                age = int(float(mc_age))
        except (ValueError, TypeError):
            pass
        if age is None and birthday:
            try:
                bd = datetime.strptime(birthday, "%d/%m/%Y").date()
                age = TODAY.year - bd.year - ((TODAY.month, TODAY.day) < (bd.month, bd.day))
            except: pass

        # Tenure from join date
        tenure = None
        if join_year:
            tenure = TODAY.year - join_year

        # Handicap
        hcp_raw = r[COLS["handicap"]]
        handicap = None
        if hcp_raw is not None and str(hcp_raw).strip() not in ("","None"):
            try: handicap = round(float(hcp_raw), 1)
            except: pass

        members.append({
            "id":              make_id(first, last, g("ggs_id"), g("member_code")),
            "status":          status,
            "first":           first,
            "last":            last,
            "email":           email,
            "phone":           phone,
            "gender":          gender,
            "birthday":        birthday,
            "age":             age,
            "age_group":       age_group(age),
            "join_date":       join_date,
            "join_year":       join_year,
            "tenure":          tenure,
            "tenure_bucket":   tenure_bucket(tenure),
            "suffix":          "",
            # MiClub-specific
            "member_code":     g("member_code"),
            "membership_type": g("membership_type"),
            "handicap":        handicap,
            "city":            g("city"),
            # GG equipment
            "driver":          g("driver"),
            "woods":           g("woods"),
            "irons":           g("irons"),
            "wedges":          g("wedges"),
            "putter":          g("putter"),
            "ball":            g("ball_brand"),
            "glove_brand":     g("glove_brand"),
            "glove_size":      g("glove_size"),
            "apparel_brand":   g("apparel_brand"),
            "apparel_size":    g("apparel_size"),
            "shoe_brand":      g("shoe_brand"),
            "shoe_size":       g("shoe_size"),
        })

    # Sort: matched first (alphabetical), then MC-only, then GG-only
    ORDER = {"Matched": 0, "Missing from Golf Genius": 1, "Missing from MiClub": 2}
    members.sort(key=lambda m: (ORDER.get(m["status"], 9), m["last"].lower(), m["first"].lower()))

    payload = {"members": members, "generated": TODAY.isoformat()}
    js = f"window.MEMBER_LOOKUP_DATA = {json.dumps(payload, separators=(',', ':'))};"

    for out in OUTPUTS:
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(js, encoding="utf-8")
        kb = out.stat().st_size / 1024
        print(f"  Wrote {out.relative_to(BASE)}  ({kb:.0f} KB, {len(members)} members)")

    mc_only = sum(1 for m in members if m["status"] == "Missing from Golf Genius")
    gg_only = sum(1 for m in members if m["status"] == "Missing from MiClub")
    matched = sum(1 for m in members if m["status"] == "Matched")
    print(f"\nSummary: {matched} matched | {mc_only} MiClub-only | {gg_only} GG-only")


if __name__ == "__main__":
    main()
